#!/usr/bin/env python3
"""
Sales Comparison Script
Compares Amazon sales data between two business reports (Older vs Newer)
Outputs result as JSON for web consumption
"""

import pandas as pd
import sys
import json
import argparse
from decimal import Decimal
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

def parse_currency(value):
    """Parse currency string to float"""
    if pd.isna(value) or value == '':
        return 0.0
    # Remove £ symbol, commas, and convert to float
    cleaned = str(value).replace('£', '').replace(',', '').strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def parse_percentage(value):
    """Parse percentage string to float"""
    if pd.isna(value) or value == '':
        return 0.0
    # Remove % symbol and convert to float
    cleaned = str(value).replace('%', '').strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def parse_number(value):
    """Parse numeric string with commas to float"""
    if pd.isna(value) or value == '':
        return 0.0
    # Remove commas and convert to float
    cleaned = str(value).replace(',', '').strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def load_and_clean_data(filepath):
    """Load CSV and clean the data"""
    try:
        df = pd.read_csv(filepath)
        
        # Clean column names (remove extra spaces)
        df.columns = df.columns.str.strip()
        
        # Parse currency columns
        currency_columns = ['Ordered Product Sales', 'Ordered product sales – B2B']
        for col in currency_columns:
            if col in df.columns:
                df[col] = df[col].apply(parse_currency)
        
        # Parse percentage columns
        percentage_columns = [
            'Session percentage – Total', 'Session percentage – Total – B2B',
            'Page views percentage – Total', 'Page views percentage – Total – B2B',
            'Featured Offer (Buy Box) percentage', 'Featured Offer (Buy Box) percentage – B2B',
            'Unit Session Percentage', 'Unit session percentage – B2B'
        ]
        for col in percentage_columns:
            if col in df.columns:
                df[col] = df[col].apply(parse_percentage)
        
        # Ensure numeric columns are properly typed
        numeric_columns = [
            'Sessions – Total', 'Sessions – Total – B2B',
            'Page views – Total', 'Page views – Total – B2B',
            'Units ordered', 'Units ordered – B2B',
            'Total order items', 'Total order items – B2B'
        ]
        for col in numeric_columns:
            if col in df.columns:
                # Use apply(parse_number) instead of pd.to_numeric directly to handle commas
                df[col] = df[col].apply(parse_number)
        
        return df
    
    except Exception as e:
        # Return generic error instead of printing, or let it propagate
        raise Exception(f"Error loading {filepath}: {e}")

def compare_reports(old_df, new_df):
    """Compare sales data between two reports"""
    
    # Handle duplicate SKUs by aggregating numeric data and keeping first non-numeric
    # This ensures we don't lose sales data when multiple rows share a SKU
    
    # Process Old Report
    if 'SKU' in old_df.columns:
        numeric_old = old_df.groupby('SKU').sum(numeric_only=True)
        # Get metadata (Title, ASIN, etc.) by taking the first occurrence
        meta_old = old_df.groupby('SKU').first(numeric_only=False)
        # Combine: meta_old has the structure, update numeric columns with sums
        old_df = meta_old
        old_df.update(numeric_old)
    else:
        # Fallback if SKU column missing (shouldn't happen based on parser)
        old_df = old_df.set_index('SKU') if 'SKU' in old_df.columns else old_df

    # Process New Report
    if 'SKU' in new_df.columns:
        numeric_new = new_df.groupby('SKU').sum(numeric_only=True)
        meta_new = new_df.groupby('SKU').first(numeric_only=False)
        new_df = meta_new
        new_df.update(numeric_new)
    else:
        new_df = new_df.set_index('SKU') if 'SKU' in new_df.columns else new_df
    
    comparison_results = []
    
    # Get all unique SKUs from both weeks
    all_skus = set(old_df.index) | set(new_df.index)
    
    for sku in all_skus:
        # Skip if SKU is NaN
        if pd.isna(sku):
            continue
            
        result = {'SKU': sku}
        
        # Get data for each week (default to None if not present)
        old_data = None
        new_data = None

        if sku in old_df.index:
            old_data = old_df.loc[sku]
            # Handle multiple rows with same SKU - take first one
            if isinstance(old_data, pd.DataFrame):
                old_data = old_data.iloc[0]
                
        if sku in new_df.index:
            new_data = new_df.loc[sku]
            # Handle multiple rows with same SKU - take first one
            if isinstance(new_data, pd.DataFrame):
                new_data = new_data.iloc[0]
        
        # Product title (prefer New, fallback to Old)
        if new_data is not None:
            result['Product_Title'] = str(new_data.get('Title', 'Unknown'))
            result['ASIN'] = str(new_data.get('(Child) ASIN', 'Unknown'))
        elif old_data is not None:
            result['Product_Title'] = str(old_data.get('Title', 'Unknown'))
            result['ASIN'] = str(old_data.get('(Child) ASIN', 'Unknown'))
        else:
            result['Product_Title'] = 'Unknown'
            result['ASIN'] = 'Unknown'
        
        # Sales data comparison
        old_sales = float(old_data['Ordered Product Sales']) if old_data is not None and 'Ordered Product Sales' in old_data else 0.0
        new_sales = float(new_data['Ordered Product Sales']) if new_data is not None and 'Ordered Product Sales' in new_data else 0.0
        
        result['Old_Sales'] = old_sales
        result['New_Sales'] = new_sales
        result['Sales_Change'] = new_sales - old_sales
        
        # Avoid division by zero
        if old_sales > 0:
            result['Sales_Change_Percent'] = (new_sales - old_sales) / old_sales * 100
        elif new_sales > 0:
            result['Sales_Change_Percent'] = 100.0  # New product or first sales
        else:
            result['Sales_Change_Percent'] = 0.0
        
        # Units ordered comparison
        old_units = float(old_data['Units ordered']) if old_data is not None and 'Units ordered' in old_data else 0.0
        new_units = float(new_data['Units ordered']) if new_data is not None and 'Units ordered' in new_data else 0.0
        
        result['Old_Units'] = old_units
        result['New_Units'] = new_units
        result['Units_Change'] = new_units - old_units
        if old_units > 0:
            result['Units_Change_Percent'] = (new_units - old_units) / old_units * 100
        elif new_units > 0:
            result['Units_Change_Percent'] = 100.0
        else:
            result['Units_Change_Percent'] = 0.0
            
        # Sessions comparison
        old_sessions = float(old_data['Sessions – Total']) if old_data is not None and 'Sessions – Total' in old_data else 0.0
        new_sessions = float(new_data['Sessions – Total']) if new_data is not None and 'Sessions – Total' in new_data else 0.0
        
        result['Old_Sessions'] = old_sessions
        result['New_Sessions'] = new_sessions
        result['Sessions_Change'] = new_sessions - old_sessions
        
        # Page views comparison
        old_page_views = float(old_data['Page views – Total']) if old_data is not None and 'Page views – Total' in old_data else 0.0
        new_page_views = float(new_data['Page views – Total']) if new_data is not None and 'Page views – Total' in new_data else 0.0
        
        result['Old_Page_Views'] = old_page_views
        result['New_Page_Views'] = new_page_views
        result['Page_Views_Change'] = new_page_views - old_page_views
        
        # Unit Session Percentage (Conversion Rate) comparison
        old_conversion = float(old_data['Unit Session Percentage']) if old_data is not None and 'Unit Session Percentage' in old_data else 0.0
        new_conversion = float(new_data['Unit Session Percentage']) if new_data is not None and 'Unit Session Percentage' in new_data else 0.0
        
        result['Old_Conversion'] = old_conversion
        result['New_Conversion'] = new_conversion
        result['Conversion_Change'] = new_conversion - old_conversion
        
        # Buy Box percentage comparison
        old_buybox = float(old_data['Featured Offer (Buy Box) percentage']) if old_data is not None and 'Featured Offer (Buy Box) percentage' in old_data else 0.0
        new_buybox = float(new_data['Featured Offer (Buy Box) percentage']) if new_data is not None and 'Featured Offer (Buy Box) percentage' in new_data else 0.0
        
        result['Old_BuyBox'] = old_buybox
        result['New_BuyBox'] = new_buybox
        result['BuyBox_Change'] = new_buybox - old_buybox

        # Status classification
        if old_sales == 0 and new_sales > 0:
            result['Status'] = 'NO_PREV_SALES'
        elif old_sales > 0 and new_sales == 0:
            result['Status'] = 'DISCONTINUED' # Or just no sales this week
        elif result['Sales_Change_Percent'] > 20:
            result['Status'] = 'INCREASE'
        elif result['Sales_Change_Percent'] < -20:
            result['Status'] = 'DECREASE'
        else:
            result['Status'] = 'STABLE'
        
        comparison_results.append(result)
    
    return pd.DataFrame(comparison_results)

def main():
    parser = argparse.ArgumentParser(description='Compare two Amazon business reports')
    parser.add_argument('old_report', help='Path to the older report CSV')
    parser.add_argument('new_report', help='Path to the newer report CSV')
    parser.add_argument('--output-excel', help='Path to save Excel report', required=False)
    
    args = parser.parse_args()
    
    try:
        old_df = load_and_clean_data(args.old_report)
        new_df = load_and_clean_data(args.new_report)
        
        if old_df is None or new_df is None:
            print(json.dumps({"error": "Failed to load reports"}))
            sys.exit(1)
            
        comparison_df = compare_reports(old_df, new_df)
        
        # Sort by New Sales Descending (Highest sales first) for the main list
        comparison_df = comparison_df.sort_values(by='New_Sales', ascending=False)
        
        # Calculate summary stats
        total_old_sales = float(comparison_df['Old_Sales'].sum())
        total_new_sales = float(comparison_df['New_Sales'].sum())
        total_change = total_new_sales - total_old_sales
        total_change_percent = (total_change / total_old_sales * 100) if total_old_sales > 0 else 0
        
        # --- Top Line Analysis ---

        # 1. Biggest movers in sales increases (Absolute Sales Change)
        top_sales_increases = comparison_df.sort_values('Sales_Change', ascending=False).head(5)[['Product_Title', 'SKU', 'Sales_Change', 'Sales_Change_Percent']]
        top_sales_increases_dict = top_sales_increases.to_dict(orient='records')
        
        # 2. Biggest movers in sales drops (Absolute Sales Change)
        top_sales_decreases = comparison_df.sort_values('Sales_Change', ascending=True).head(5)[['Product_Title', 'SKU', 'Sales_Change', 'Sales_Change_Percent']]
        top_sales_decreases_dict = top_sales_decreases.to_dict(orient='records')

        # 3. Biggest buy box changes (Absolute Change) - Increases
        top_buybox_increases = comparison_df.sort_values('BuyBox_Change', ascending=False).head(5)[['Product_Title', 'SKU', 'BuyBox_Change', 'New_BuyBox']]
        top_buybox_increases_dict = top_buybox_increases.to_dict(orient='records')

        # 4. Biggest buy box changes (Absolute Change) - Decreases
        top_buybox_decreases = comparison_df.sort_values('BuyBox_Change', ascending=True).head(5)[['Product_Title', 'SKU', 'BuyBox_Change', 'New_BuyBox']]
        top_buybox_decreases_dict = top_buybox_decreases.to_dict(orient='records')

        # 5. Loss in sales due to buy box (Buy Box Drop AND Sales Drop)
        # Filter where Buy Box Change < -3% AND Sales Change < 0
        buybox_sales_loss_mask = (comparison_df['BuyBox_Change'] < -3) & (comparison_df['Sales_Change'] < 0)
        buybox_sales_loss_df = comparison_df[buybox_sales_loss_mask].sort_values('Sales_Change', ascending=True).head(5)
        
        if not buybox_sales_loss_df.empty:
            buybox_sales_loss_data = buybox_sales_loss_df[['Product_Title', 'SKU', 'Sales_Change', 'BuyBox_Change']]
            buybox_sales_loss_dict = buybox_sales_loss_data.to_dict(orient='records')
        else:
            buybox_sales_loss_data = pd.DataFrame()
            buybox_sales_loss_dict = []

        # 6. Gain in sales due to buy box (Buy Box Increase AND Sales Increase)
        # Filter where Buy Box Change > 3% AND Sales Change > 0
        buybox_sales_gain_mask = (comparison_df['BuyBox_Change'] > 3) & (comparison_df['Sales_Change'] > 0)
        buybox_sales_gain_df = comparison_df[buybox_sales_gain_mask].sort_values('Sales_Change', ascending=False).head(5)
        
        if not buybox_sales_gain_df.empty:
            buybox_sales_gain_data = buybox_sales_gain_df[['Product_Title', 'SKU', 'Sales_Change', 'BuyBox_Change']]
            buybox_sales_gain_dict = buybox_sales_gain_data.to_dict(orient='records')
        else:
            buybox_sales_gain_data = pd.DataFrame()
            buybox_sales_gain_dict = []

        if args.output_excel:
            with pd.ExcelWriter(args.output_excel, engine='openpyxl') as writer:
                # Tab 1: Top Line Analysis
                
                # Write Summary Stats
                summary_data = {
                   'Metric': ['Previous Total Sales', 'Current Total Sales', 'Net Change', '% Change'],
                   'Value': [total_old_sales, total_new_sales, total_change, total_change_percent]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Top Line Analysis', startrow=0, index=False)
                
                row_offset = 6
                
                # Sales Increases
                pd.DataFrame({'Section': ['Top Sales Movers - Biggest Increases']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset, index=False, header=False)
                top_sales_increases.to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)
                row_offset += len(top_sales_increases) + 4
                
                # Sales Decreases
                pd.DataFrame({'Section': ['Top Sales Movers - Biggest Drops']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset, index=False, header=False)
                top_sales_decreases.to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)
                row_offset += len(top_sales_decreases) + 4
                
                # Buy Box Increases
                pd.DataFrame({'Section': ['Buy Box Stability - Biggest Gains']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset, index=False, header=False)
                top_buybox_increases.to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)
                row_offset += len(top_buybox_increases) + 4
                
                # Buy Box Decreases
                pd.DataFrame({'Section': ['Buy Box Stability - Biggest Drops']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset, index=False, header=False)
                top_buybox_decreases.to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)
                row_offset += len(top_buybox_decreases) + 4
                
                # Buy Box Sales Impact - Gain
                pd.DataFrame({'Section': ['Buy Box Sales Impact - Sales Gained (BB Increase)']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset, index=False, header=False)
                if not buybox_sales_gain_data.empty:
                    buybox_sales_gain_data.to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)
                    row_offset += len(buybox_sales_gain_data) + 4
                else:
                    pd.DataFrame({'Message': ['No significant gains']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)
                    row_offset += 4
                    
                # Buy Box Sales Impact - Loss
                pd.DataFrame({'Section': ['Buy Box Sales Impact - Sales Lost (BB Decrease)']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset, index=False, header=False)
                if not buybox_sales_loss_data.empty:
                    buybox_sales_loss_data.to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)
                else:
                    pd.DataFrame({'Message': ['No significant losses']}).to_excel(writer, sheet_name='Top Line Analysis', startrow=row_offset+1, index=False)

                # Tab 2: Product Performance Review
                comparison_df.to_excel(writer, sheet_name='Product Performance Review', index=False)

                # --- Formatting ---
                from openpyxl.styles import Font, Alignment
                
                # 1. Format Top Line Analysis
                if 'Top Line Analysis' in writer.sheets:
                    ws1 = writer.sheets['Top Line Analysis']
                    ws1.column_dimensions['A'].width = 40
                    ws1.column_dimensions['B'].width = 20
                    ws1.column_dimensions['C'].width = 15
                    ws1.column_dimensions['D'].width = 20

                    # Summary Table formatting (Rows 2-5)
                    for row_idx in range(2, 6): # 1-based indices 2,3,4,5
                        metric_cell = ws1.cell(row=row_idx, column=1)
                        value_cell = ws1.cell(row=row_idx, column=2)
                        metric_val = metric_cell.value
                        
                        if metric_val and isinstance(metric_val, str):
                            if 'Sales' in metric_val or 'Change' in metric_val and '%' not in metric_val:
                                if isinstance(value_cell.value, (int, float)):
                                    value_cell.number_format = '#,##0.00'
                            elif '%' in metric_val:
                                if isinstance(value_cell.value, (int, float)):
                                    value_cell.value = value_cell.value / 100
                                    value_cell.number_format = '0.00%'

                    # Find all table headers to format sections
                    # Iterate rows to find headers and apply formatting to data below
                    for row in ws1.iter_rows():
                        # Bold section headers
                        if row[0].value and isinstance(row[0].value, str) and ("Top" in row[0].value or "Buy Box" in row[0].value):
                            row[0].font = Font(bold=True, size=12)
                        
                        # Bold table headers
                        if row[1].value == 'SKU':
                            for cell in row:
                                if cell.value: cell.font = Font(bold=True)
                                
                            # If this is a header row, we can identify data columns
                            # Col 3 and 4 (indices 2, 3)
                            col_c_name = row[2].value
                            col_d_name = row[3].value
                            
                            format_c = None
                            format_d = '0.00%' # Default for last col
                            
                            if 'Sales_Change' in str(col_c_name):
                                format_c = '#,##0.00'
                            elif 'BuyBox_Change' in str(col_c_name):
                                format_c = '0.00' # Points change
                            
                            # Identify this row's index to start formatting data below it
                            header_row_idx = row[0].row # 1-based
                            
                            # Apply to subsequent rows until empty
                            current_row = header_row_idx + 1
                            while current_row <= ws1.max_row:
                                cell_sku = ws1.cell(row=current_row, column=2)
                                if not cell_sku.value: 
                                    break
                                
                                # Format Col 3
                                if format_c:
                                    cell = ws1.cell(row=current_row, column=3)
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = format_c
                                
                                # Format Col 4
                                if format_d:
                                    cell = ws1.cell(row=current_row, column=4)
                                    if isinstance(cell.value, (int, float)):
                                        cell.value = cell.value / 100
                                        cell.number_format = format_d
                                current_row += 1


                # 2. Format Product Performance Review
                if 'Product Performance Review' in writer.sheets:
                    ws2 = writer.sheets['Product Performance Review']
                    
                    # Set column widths
                    ws2.column_dimensions['A'].width = 15 # SKU
                    ws2.column_dimensions['B'].width = 40 # Product Title
                    ws2.column_dimensions['C'].width = 15 # ASIN
                    ws2.column_dimensions['D'].width = 15 # Old Sales
                    ws2.column_dimensions['E'].width = 15 # New Sales
                    ws2.column_dimensions['F'].width = 15 # Sales Change
                    ws2.column_dimensions['G'].width = 12 # % Change
                    
                    # Map columns
                    header_row = next(ws2.iter_rows(min_row=1, max_row=1))
                    col_map = {cell.value: i+1 for i, cell in enumerate(header_row)}
                    
                    currency_cols = ['Old_Sales', 'New_Sales', 'Sales_Change']
                    percent_cols = ['Sales_Change_Percent', 'Old_Conversion', 'New_Conversion', 'Conversion_Change', 
                                    'Old_BuyBox', 'New_BuyBox', 'BuyBox_Change', 'Units_Change_Percent']
                    
                    # Bold header
                    for cell in header_row:
                        cell.font = Font(bold=True)
                    
                    # Apply formatting column by column
                    for col_name, col_idx in col_map.items():
                        if col_name in currency_cols:
                            for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                                cell = row[0]
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0.00'
                        elif col_name in percent_cols:
                            for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                                cell = row[0]
                                if isinstance(cell.value, (int, float)):
                                    cell.value = cell.value / 100
                                    cell.number_format = '0.00%'

        # Convert to dictionary for JSON output
        # orient='records' turns it into a list of dicts
        products_data = comparison_df.to_dict(orient='records')
        
        output = {
            "summary": {
                "total_old_sales": total_old_sales,
                "total_new_sales": total_new_sales,
                "total_change": total_change,
                "total_change_percent": total_change_percent,
                "product_count": len(comparison_df)
            },
            "top_movers": {
                "sales_increases": top_sales_increases_dict, # List of top 5 winners
                "sales_decreases": top_sales_decreases_dict, # List of top 5 losers
                "buybox_increases": top_buybox_increases_dict,
                "buybox_decreases": top_buybox_decreases_dict,
                "buybox_sales_impact": buybox_sales_loss_dict,
                "buybox_sales_gain": buybox_sales_gain_dict
            },
            "products": products_data
        }
        
        print(json.dumps(output))
        
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main()
